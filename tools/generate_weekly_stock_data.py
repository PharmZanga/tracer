import json
import os
import sys
from pathlib import Path

import openpyxl


ROOT = Path(r"C:\Users\Zanga Musakuzi\Desktop\zammsa folder\weekly inventory emms stock status")
OUT = Path(__file__).resolve().parents[1] / "src" / "weeklyStockData.js"

SOURCES = [
    {
        "file": ROOT / "may zammsa emms stock status" / "EMMS stock Postion  As at 1st May 2026.xlsx",
        "periods": [
            {"sheet": "Sheet1", "date": "2026-05-01", "label": "1 May 2026", "stream": "EMMS"},
        ],
    },
    {
        "file": ROOT / "may zammsa emms stock status" / "Stock Position as at 8th May 2026.xlsx",
        "periods": [
            {"sheet": "EMMS", "date": "2026-05-08", "label": "8 May 2026", "stream": "EMMS"},
            {"sheet": "LAB", "date": "2026-05-08", "label": "8 May 2026", "stream": "LAB"},
        ],
    },
    {
        "file": ROOT / "may zammsa emms stock status" / "STOCK POSITION AS AT 15TH MAY 2026.xlsx",
        "periods": [
            {"sheet": "EMMS", "date": "2026-05-15", "label": "15 May 2026", "stream": "EMMS"},
            {"sheet": "LAB", "date": "2026-05-15", "label": "15 May 2026", "stream": "LAB"},
        ],
    },
    {
        "file": ROOT / "may zammsa emms stock status" / "ZAMMSA STOCK POSITION.xlsx",
        "periods": [
            {"sheet": "EMMS 22MAY", "date": "2026-05-22", "label": "22 May 2026", "stream": "EMMS"},
            {"sheet": "LAB 22MA", "date": "2026-05-22", "label": "22 May 2026", "stream": "LAB"},
            {"sheet": "EMMS 29TH MAY", "date": "2026-05-29", "label": "29 May 2026", "stream": "EMMS"},
            {"sheet": "LAV 29MAY", "date": "2026-05-29", "label": "29 May 2026", "stream": "LAB"},
            {"sheet": "EMMS 5JUNE", "date": "2026-06-05", "label": "5 June 2026", "stream": "EMMS"},
            {"sheet": "LAB 5JUNE", "date": "2026-06-05", "label": "5 June 2026", "stream": "LAB"},
        ],
    },
    {
        "file": ROOT / "june" / "Stock Position 13 and 19 june 2026.xlsx",
        "periods": [
            {"sheet": "EMMS-13 JUNE", "date": "2026-06-13", "label": "13 June 2026", "stream": "EMMS"},
            {"sheet": "LAB-13JUNE", "date": "2026-06-13", "label": "13 June 2026", "stream": "LAB"},
            {"sheet": "EMMS-19 JUNE", "date": "2026-06-19", "label": "19 June 2026", "stream": "EMMS"},
            {"sheet": "LAB-19JUNE", "date": "2026-06-19", "label": "19 June 2026", "stream": "LAB"},
        ],
    },
    {
        "file": ROOT / "june" / "STOCK POSITION 19 AND 26 JUNE 2026.xlsx",
        "periods": [
            {"sheet": "26June", "date": "2026-06-26", "label": "26 June 2026", "stream": "EMMS"},
            {"sheet": "LAB26 June", "date": "2026-06-26", "label": "26 June 2026", "stream": "LAB"},
        ],
    },
    {
        "file": ROOT / "july" / "STOCK POSITION 04-JULY 2026.xlsx",
        "periods": [
            {"sheet": "EMMS04-JULY", "date": "2026-07-04", "label": "4 July 2026", "stream": "EMMS"},
            {"sheet": "LAB 4-JULY", "date": "2026-07-04", "label": "4 July 2026", "stream": "LAB"},
        ],
    },
]


def clean(value):
    if value is None:
        return None
    if isinstance(value, str):
        value = " ".join(value.strip().split())
        return value or None
    return value


def num(value):
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def pct(value):
    value = num(value)
    if value is None:
        return 0
    if value > 1:
        value = value / 100
    return max(0, min(value, 1))


def header_text(value):
    return str(value or "").strip().lower()


def find_overall(ws):
    for row in range(1, 12):
        for col in range(1, 8):
            if "overall" in header_text(ws.cell(row, col).value):
                for offset in range(1, 4):
                    value = num(ws.cell(row, col + offset).value)
                    if value is not None:
                        return pct(value)
    for cell in ("C3", "B3", "C2", "B2"):
        value = num(ws[cell].value)
        if value is not None:
            return pct(value)
    return 0


def find_layout(ws):
    category = None
    item = None
    for row in range(1, 45):
        for col in range(1, 9):
            current = header_text(ws.cell(row, col).value)
            nxt = header_text(ws.cell(row, col + 1).value)
            if current in {"product category", "category"} and ("availability" in nxt or "stock availability" in nxt):
                if category is None:
                    category = {"row": row + 1, "name_col": col, "availability_col": col + 1}
                else:
                    item = {"row": row + 1, "index_col": col - 1, "name_col": col, "availability_col": col + 1}
                    return category, item
    if category is None:
        category = {"row": 18, "name_col": 2, "availability_col": 3}
    if item is None:
        item = {"row": category["row"], "index_col": 5, "name_col": 6, "availability_col": 7}
    return category, item


def is_item_index(value):
    if isinstance(value, (int, float)):
        return True
    if isinstance(value, str) and value.strip().upper().startswith("#REF"):
        return True
    return False


def parse_sheet(ws, config, source_name):
    category_layout, item_layout = find_layout(ws)
    start_row = min(category_layout["row"], item_layout["row"])
    end_row = min(ws.max_row, 1100)
    categories = []
    items = []
    current_category = None
    blank_streak = 0

    for row in range(start_row, end_row + 1):
        category_name = clean(ws.cell(row, category_layout["name_col"]).value)
        category_availability = num(ws.cell(row, category_layout["availability_col"]).value)
        item_index = ws.cell(row, item_layout["index_col"]).value
        item_name = clean(ws.cell(row, item_layout["name_col"]).value)
        item_availability = num(ws.cell(row, item_layout["availability_col"]).value)

        has_data = any(value not in (None, "") for value in (category_name, category_availability, item_index, item_name, item_availability))
        blank_streak = 0 if has_data else blank_streak + 1
        if blank_streak > 30:
            break

        if category_name and category_availability is not None and header_text(category_name) not in {"product category", "category"}:
            categories.append({
                "name": category_name,
                "availability": round(pct(category_availability), 4),
            })

        if item_name and item_availability is not None:
            if is_item_index(item_index):
                items.append({
                    "category": current_category or "Uncategorised",
                    "name": item_name,
                    "availability": round(pct(item_availability), 4),
                    "status": "Available" if pct(item_availability) > 0 else "Stockout",
                })
            elif header_text(item_name) not in {"category"}:
                current_category = item_name

    categories = dedupe_rows(categories, "name")
    items = dedupe_rows(items, "name")
    categories.sort(key=lambda item: (item["availability"], item["name"]))
    items.sort(key=lambda item: (item["availability"], item["category"], item["name"]))
    available_items = sum(1 for item in items if item["availability"] > 0)
    return {
        "id": f'{config["stream"].lower()}-{config["date"]}',
        "date": config["date"],
        "label": config["label"],
        "stream": config["stream"],
        "source": source_name,
        "overallAvailability": round(find_overall(ws), 4),
        "counts": {
            "categories": len(categories),
            "items": len(items),
            "availableItems": available_items,
            "stockoutItems": len(items) - available_items,
        },
        "categories": categories,
        "items": items,
    }


def dedupe_rows(rows, key):
    deduped = {}
    for row in rows:
        deduped[clean(row.get(key))] = row
    return list(deduped.values())


def main():
    periods = []
    seen = set()
    for source in SOURCES:
        path = source["file"]
        if not path.exists():
            raise FileNotFoundError(path)
        workbook = openpyxl.load_workbook(path, read_only=False, data_only=True, keep_links=False)
        try:
            for config in source["periods"]:
                key = (config["stream"], config["date"])
                if key in seen:
                    continue
                if config["sheet"] not in workbook.sheetnames:
                    raise KeyError(f'{path.name}: missing sheet {config["sheet"]}')
                periods.append(parse_sheet(workbook[config["sheet"]], config, path.name))
                seen.add(key)
        finally:
            workbook.close()

    periods.sort(key=lambda period: (period["date"], period["stream"]))
    output = (
        "export const weeklyStockPeriods = "
        + json.dumps(periods, indent=2)
        + ";\n\nexport const latestWeeklyStock = weeklyStockPeriods.at(-1);\n"
    )
    OUT.write_text(output, encoding="utf-8")
    for period in periods:
        counts = period["counts"]
        print(period["stream"], period["label"], period["overallAvailability"], counts)
    sys.stdout.flush()
    os._exit(0)


if __name__ == "__main__":
    main()
