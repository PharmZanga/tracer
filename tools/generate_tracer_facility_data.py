import json
import math
import re
from collections import defaultdict
from pathlib import Path

import openpyxl


JULY_WEEK4_CLEAN_WORKBOOK = Path(r"C:\Users\Zanga Musakuzi\Desktop\NSCCU DATA ANALYSIS\PROVINCIAL  tracer SUBMISSION\tracer summery report clean data\july\28.07.26 summary reports.xlsx")
JULY_WEEK5_CLEAN_WORKBOOK = Path(r"C:\Users\Zanga Musakuzi\Desktop\NSCCU DATA ANALYSIS\PROVINCIAL  tracer SUBMISSION\tracer summery report clean data\july\TRACER SUMMARY 02 AUAGUST 2026.xlsx")
AUGUST_WEEK1_CLEAN_WORKBOOK = Path(r"C:\Users\Zanga Musakuzi\Desktop\NSCCU DATA ANALYSIS\PROVINCIAL  tracer SUBMISSION\tracer summery report clean data\august\week 1\9.8.2026 tracer summary.xlsx")
AUGUST_WEEK2_CLEAN_WORKBOOK = Path(r"C:\Users\Zanga Musakuzi\Desktop\NSCCU DATA ANALYSIS\PROVINCIAL  tracer SUBMISSION\tracer summery report clean data\august\week 2\tracer summary 16-08-2026.xlsx")
AUGUST_WEEK3_CLEAN_WORKBOOK = Path(r"C:\Users\Zanga Musakuzi\Desktop\NSCCU DATA ANALYSIS\PROVINCIAL  tracer SUBMISSION\tracer summery report clean data\august\week 3\23.08.2026.xlsx")


WORKBOOKS = [
    {
        "path": Path(r"C:\Users\Zanga Musakuzi\Desktop\NSCCU DATA ANALYSIS\PROVINCIAL  tracer SUBMISSION\tracer summery report clean data\feb\TRACER SUMMARY REPORTS  JANUARY 2026 (3).xlsx"),
        "sheet": "4-01-2026",
        "reportDate": "2026-01-04",
        "label": "Week 1 - 4 January 2026",
        "month": "2026-01",
        "week": "Week 1",
    },
    {
        "path": Path(r"C:\Users\Zanga Musakuzi\Desktop\NSCCU DATA ANALYSIS\PROVINCIAL  tracer SUBMISSION\tracer summery report clean data\feb\TRACER SUMMARY REPORTS  JANUARY 2026 (3).xlsx"),
        "sheet": "11-01-2026",
        "reportDate": "2026-01-11",
        "label": "Week 2 - 11 January 2026",
        "month": "2026-01",
        "week": "Week 2",
    },
    {
        "path": Path(r"C:\Users\Zanga Musakuzi\Desktop\NSCCU DATA ANALYSIS\PROVINCIAL  tracer SUBMISSION\tracer summery report clean data\feb\TRACER SUMMARY REPORTS  JANUARY 2026 (3).xlsx"),
        "sheet": "18-01-2026",
        "reportDate": "2026-01-18",
        "label": "Week 3 - 18 January 2026",
        "month": "2026-01",
        "week": "Week 3",
    },
    {
        "path": Path(r"C:\Users\Zanga Musakuzi\Desktop\NSCCU DATA ANALYSIS\PROVINCIAL  tracer SUBMISSION\tracer summery report clean data\feb\TRACER SUMMARY REPORTS  JANUARY 2026 (3).xlsx"),
        "sheet": "25-01-2026",
        "reportDate": "2026-01-25",
        "label": "Week 4 - 25 January 2026",
        "month": "2026-01",
        "week": "Week 4",
    },
    {
        "path": Path(r"C:\Users\Zanga Musakuzi\Desktop\NSCCU DATA ANALYSIS\PROVINCIAL  tracer SUBMISSION\tracer summery report clean data\feb\TRACER SUMMARY REPORTS  JANUARY 2026 (3).xlsx"),
        "sheet": "31-01-2026",
        "reportDate": "2026-01-31",
        "label": "Week 5 - 31 January 2026",
        "month": "2026-01",
        "week": "Week 5",
    },
    {
        "rawSources": [
            {"province": "MUCHINGA PROVINCE", "path": Path(r"C:\Users\Zanga Musakuzi\Desktop\NSCCU DATA ANALYSIS\PROVINCIAL  tracer SUBMISSION\province submissions\February province submission\week 1 feb\06.02.26 MUCHINGA 2026 TRACER WEEKLY REPORT.xlsx")},
            {"province": "NORTHERN PROVINCE", "path": Path(r"C:\Users\Zanga Musakuzi\Desktop\NSCCU DATA ANALYSIS\PROVINCIAL  tracer SUBMISSION\province submissions\February province submission\week 1 feb\07.02.26 NORTHERN PROVINCE 2024 TRACER WEEKLY REPORT PROVINCES.xlsx")},
            {"province": "CENTRAL PROVINCE", "path": Path(r"C:\Users\Zanga Musakuzi\Desktop\NSCCU DATA ANALYSIS\PROVINCIAL  tracer SUBMISSION\province submissions\February province submission\week 1 feb\7_2_2025 CENTRAL PROVINCE 2025 TRACER WEEKLY REPORT.xlsx")},
            {"province": "EASTERN PROVINCE", "path": Path(r"C:\Users\Zanga Musakuzi\Desktop\NSCCU DATA ANALYSIS\PROVINCIAL  tracer SUBMISSION\province submissions\February province submission\week 1 feb\7TH FEB EASTERN PROVINCE 2026 TRACER WEEKLY REPORT PROVINCES (4).xlsx")},
            {"province": "COPPERBELT PROVINCE", "path": Path(r"C:\Users\Zanga Musakuzi\Desktop\NSCCU DATA ANALYSIS\PROVINCIAL  tracer SUBMISSION\province submissions\February province submission\week 1 feb\08.02.26 COPPERBELT PROVINCE  TRACER WEEKLY REPORT PROVINCES.xlsx")},
            {"province": "NORTH-WESTERN PROVINCE", "path": Path(r"C:\Users\Zanga Musakuzi\Desktop\NSCCU DATA ANALYSIS\PROVINCIAL  tracer SUBMISSION\province submissions\February province submission\week 1 feb\08-02-2026 NORTHWESTERN TRACER WEEKLY REPORT PROVINCES.xlsx")},
            {"province": "WESTERN PROVINCE", "path": Path(r"C:\Users\Zanga Musakuzi\Desktop\NSCCU DATA ANALYSIS\PROVINCIAL  tracer SUBMISSION\province submissions\February province submission\week 1 feb\08-02-2026 WESTERN PROVINCE 2025 TRACER WEEKLY REPORT.xlsx")},
            {"province": "LUAPULA PROVINCE", "path": Path(r"C:\Users\Zanga Musakuzi\Desktop\NSCCU DATA ANALYSIS\PROVINCIAL  tracer SUBMISSION\province submissions\February province submission\week 1 feb\LUAPULA PROVINCE 2025 TRACER WEEKLY REPORT 7 2 26.xlsx")},
            {"province": "LUSAKA PROVINCE", "path": Path(r"C:\Users\Zanga Musakuzi\Desktop\NSCCU DATA ANALYSIS\PROVINCIAL  tracer SUBMISSION\province submissions\February province submission\week 1 feb\LUSAKA PROVINCE TRACER WEEKLY REPORT PROVINCES-06.02.2026.xlsx")},
            {"province": "SOUTHERN PROVINCE", "path": Path(r"C:\Users\Zanga Musakuzi\Desktop\NSCCU DATA ANALYSIS\PROVINCIAL  tracer SUBMISSION\province submissions\February province submission\week 1 feb\SOUTHERN PROVINCE 2026 TRACER WEEKLY REPORT PROVINCES-WEEK ENDING 06.02.26 -Final.xlsx")},
        ],
        "source": "February Week 1 provincial raw submissions",
        "reportDate": "2026-02-08",
        "label": "Week 1 - 8 February 2026",
        "month": "2026-02",
        "week": "Week 1",
    },
    {
        "rawSources": [
            {"province": "MUCHINGA PROVINCE", "path": Path(r"C:\Users\Zanga Musakuzi\Desktop\NSCCU DATA ANALYSIS\PROVINCIAL  tracer SUBMISSION\province submissions\February province submission\weeek 2\13.02.2026 MUCHINGA 2026 TRACER WEEKLY REPORT (3).xlsx")},
            {"province": "NORTHERN PROVINCE", "path": Path(r"C:\Users\Zanga Musakuzi\Desktop\NSCCU DATA ANALYSIS\PROVINCIAL  tracer SUBMISSION\province submissions\February province submission\weeek 2\13.02.2026 NORTHERN PROVINCE 2024 TRACER WEEKLY REPORT PROVINCES.xlsx")},
            {"province": "CENTRAL PROVINCE", "path": Path(r"C:\Users\Zanga Musakuzi\Desktop\NSCCU DATA ANALYSIS\PROVINCIAL  tracer SUBMISSION\province submissions\February province submission\weeek 2\14_2_2026 CENTRAL PROVINCE 2025 TRACER WEEKLY REPORT.xlsx")},
            {"province": "EASTERN PROVINCE", "path": Path(r"C:\Users\Zanga Musakuzi\Desktop\NSCCU DATA ANALYSIS\PROVINCIAL  tracer SUBMISSION\province submissions\February province submission\weeek 2\EASTERN PROVINCE 2026 TRACER WEEKLY REPORT PROVINCES (5).xlsx")},
            {"province": "COPPERBELT PROVINCE", "path": Path(r"C:\Users\Zanga Musakuzi\Desktop\NSCCU DATA ANALYSIS\PROVINCIAL  tracer SUBMISSION\province submissions\February province submission\weeek 2\14.2.26 COPPERBELT PROVINCE  TRACER WEEKLY REPORT PROVINCES.xlsx")},
            {"province": "NORTH-WESTERN PROVINCE", "path": Path(r"C:\Users\Zanga Musakuzi\Desktop\NSCCU DATA ANALYSIS\PROVINCIAL  tracer SUBMISSION\province submissions\February province submission\weeek 2\14-02-2026-NORTHWESTERN TRACER WEEKLY REPORT.xlsx")},
            {"province": "WESTERN PROVINCE", "path": Path(r"C:\Users\Zanga Musakuzi\Desktop\NSCCU DATA ANALYSIS\PROVINCIAL  tracer SUBMISSION\province submissions\February province submission\weeek 2\15-02-26 WESTERN PROVINCE 2025 TRACER WEEKLY REPORT.xlsx")},
            {"province": "LUAPULA PROVINCE", "path": Path(r"C:\Users\Zanga Musakuzi\Desktop\NSCCU DATA ANALYSIS\PROVINCIAL  tracer SUBMISSION\province submissions\February province submission\weeek 2\LUAPULA PROVINCE 2025 TRACER WEEKLY REPORT 14 2 26.xlsx")},
            {"province": "LUSAKA PROVINCE", "path": Path(r"C:\Users\Zanga Musakuzi\Desktop\NSCCU DATA ANALYSIS\PROVINCIAL  tracer SUBMISSION\province submissions\February province submission\weeek 2\LUSAKA PROVINCE 2026 TRACER WEEKLY REPORT PROVINCES 13.02.2026.xlsx")},
            {"province": "SOUTHERN PROVINCE", "path": Path(r"C:\Users\Zanga Musakuzi\Desktop\NSCCU DATA ANALYSIS\PROVINCIAL  tracer SUBMISSION\province submissions\February province submission\weeek 2\SOUTHERN PROVINCE 2026 TRACER WEEKLY REPORT PROVINCES-WEEK ENDING 13.02.26 -FINAL.xlsx")},
        ],
        "source": "February Week 2 provincial raw submissions",
        "reportDate": "2026-02-15",
        "label": "Week 2 - 15 February 2026",
        "month": "2026-02",
        "week": "Week 2",
    },
    {
        "path": Path(r"C:\Users\Zanga Musakuzi\Desktop\NSCCU DATA ANALYSIS\PROVINCIAL  tracer SUBMISSION\tracer summery report clean data\feb\TRACER WEEKLY SUMMARY REPORTS 22.02.26.xlsx"),
        "sheet": "CV.2.22.26",
        "label": "Week 3 - 22 February 2026",
        "month": "2026-02",
        "week": "Week 3",
    },
    {
        "path": Path(r"C:\Users\Zanga Musakuzi\Desktop\NSCCU DATA ANALYSIS\PROVINCIAL  tracer SUBMISSION\tracer summery report clean data\feb\28.02.2026 SUMMARY REPORTS.xlsx"),
        "sheet": "SUMMARY REPORTS",
        "label": "Week 4 - 28 February 2026",
        "month": "2026-02",
        "week": "Week 4",
    },
    {
        "rawSources": [
            {"province": "MUCHINGA PROVINCE", "path": Path(r"C:\Users\Zanga Musakuzi\Desktop\NSCCU DATA ANALYSIS\PROVINCIAL  tracer SUBMISSION\province submissions\march province submission\week 1\9.3.26 MUCHINGA 2026 TRACER WEEKLY REPORT (5).xlsx")},
            {"province": "NORTHERN PROVINCE", "path": Path(r"C:\Users\Zanga Musakuzi\Desktop\NSCCU DATA ANALYSIS\PROVINCIAL  tracer SUBMISSION\province submissions\march province submission\week 1\06.03.26 NORTHERN PROVINCE 2024 TRACER WEEKLY REPORT PROVINCES.xlsx")},
            {"province": "CENTRAL PROVINCE", "path": Path(r"C:\Users\Zanga Musakuzi\Desktop\NSCCU DATA ANALYSIS\PROVINCIAL  tracer SUBMISSION\province submissions\march province submission\week 1\7_3_2026 CENTRAL PROVINCE 2025 TRACER WEEKLY REPORT.xlsx")},
            {"province": "EASTERN PROVINCE", "path": Path(r"C:\Users\Zanga Musakuzi\Desktop\NSCCU DATA ANALYSIS\PROVINCIAL  tracer SUBMISSION\province submissions\march province submission\week 1\7th Feb EASTERN PROVINCE 2026 TRACER WEEKLY REPORT PROVINCES (8).xlsx")},
            {"province": "COPPERBELT PROVINCE", "path": Path(r"C:\Users\Zanga Musakuzi\Desktop\NSCCU DATA ANALYSIS\PROVINCIAL  tracer SUBMISSION\province submissions\march province submission\week 1\7.3.26 COPPERBELT PROVINCE  TRACER WEEKLY REPORT PROVINCES.xlsx")},
            {"province": "NORTH-WESTERN PROVINCE", "path": Path(r"C:\Users\Zanga Musakuzi\Desktop\NSCCU DATA ANALYSIS\PROVINCIAL  tracer SUBMISSION\province submissions\march province submission\week 1\08-03-2026 NORTHWESTERN TRACER WEEKLY REPORT PROVINCES.xlsx")},
            {"province": "WESTERN PROVINCE", "path": Path(r"C:\Users\Zanga Musakuzi\Desktop\NSCCU DATA ANALYSIS\PROVINCIAL  tracer SUBMISSION\province submissions\march province submission\week 1\08-03-2026 WESTERN PROVINCE 2025 TRACER WEEKLY REPORT.xlsx")},
            {"province": "LUAPULA PROVINCE", "path": Path(r"C:\Users\Zanga Musakuzi\Desktop\NSCCU DATA ANALYSIS\PROVINCIAL  tracer SUBMISSION\province submissions\march province submission\week 1\LUAPULA PROVINCE 2025 TRACER WEEKLY REPORT 7 3 26.xlsx")},
            {"province": "LUSAKA PROVINCE", "path": Path(r"C:\Users\Zanga Musakuzi\Desktop\NSCCU DATA ANALYSIS\PROVINCIAL  tracer SUBMISSION\province submissions\march province submission\week 1\LUSAKA PROVINCE 2026 TRACER WEEKLY REPORT PROVINCES (1).xlsx")},
            {"province": "SOUTHERN PROVINCE", "path": Path(r"C:\Users\Zanga Musakuzi\Desktop\NSCCU DATA ANALYSIS\PROVINCIAL  tracer SUBMISSION\province submissions\march province submission\week 1\SOUTHERN PROVINCE 2026 TRACER WEEKLY REPORT PROVINCES-WEEK ENDING 06.03.26.xlsx")},
        ],
        "source": "March Week 1 provincial raw submissions",
        "reportDate": "2026-03-08",
        "label": "Week 1 - 8 March 2026",
        "month": "2026-03",
        "week": "Week 1",
    },
    {
        "rawSources": [
            {"province": "MUCHINGA PROVINCE", "path": Path(r"C:\Users\Zanga Musakuzi\Desktop\NSCCU DATA ANALYSIS\PROVINCIAL  tracer SUBMISSION\province submissions\march province submission\week 2\13.3.26 MUCHINGA 2026 TRACER WEEKLY REPORT.xlsx")},
            {"province": "NORTHERN PROVINCE", "path": Path(r"C:\Users\Zanga Musakuzi\Desktop\NSCCU DATA ANALYSIS\PROVINCIAL  tracer SUBMISSION\province submissions\march province submission\week 2\NORTHERN PROVINCE 2024 TRACER WEEKLY REPORT PROVINCES.xlsx")},
            {"province": "CENTRAL PROVINCE", "path": Path(r"C:\Users\Zanga Musakuzi\Desktop\NSCCU DATA ANALYSIS\PROVINCIAL  tracer SUBMISSION\province submissions\march province submission\week 2\14_3_2026 CENTRAL PROVINCE 2025 TRACER WEEKLY REPORT.xlsx")},
            {"province": "EASTERN PROVINCE", "path": Path(r"C:\Users\Zanga Musakuzi\Desktop\NSCCU DATA ANALYSIS\PROVINCIAL  tracer SUBMISSION\province submissions\march province submission\week 2\EASTERN PROVINCE 2026 TRACER WEEKLY REPORT PROVINCES (9).xlsx")},
            {"province": "COPPERBELT PROVINCE", "path": Path(r"C:\Users\Zanga Musakuzi\Desktop\NSCCU DATA ANALYSIS\PROVINCIAL  tracer SUBMISSION\province submissions\march province submission\week 2\15.3.26 COPPERBELT PROVINCE  TRACER WEEKLY REPORT PROVINCES.xlsx")},
            {"province": "NORTH-WESTERN PROVINCE", "path": Path(r"C:\Users\Zanga Musakuzi\Desktop\NSCCU DATA ANALYSIS\PROVINCIAL  tracer SUBMISSION\province submissions\march province submission\week 2\15-03-2026 NORTHWESTERN TRACER WEEKLY REPORT PROVINCES.xlsx")},
            {"province": "WESTERN PROVINCE", "path": Path(r"C:\Users\Zanga Musakuzi\Desktop\NSCCU DATA ANALYSIS\PROVINCIAL  tracer SUBMISSION\province submissions\march province submission\week 2\15-03-26 WESTERN PROVINCE 2025 TRACER WEEKLY REPORT.xlsx")},
            {"province": "LUAPULA PROVINCE", "path": Path(r"C:\Users\Zanga Musakuzi\Desktop\NSCCU DATA ANALYSIS\PROVINCIAL  tracer SUBMISSION\province submissions\march province submission\week 2\LUAPULA PROVINCE 2025 TRACER WEEKLY REPORT  15 3 25.xlsx")},
            {"province": "LUSAKA PROVINCE", "path": Path(r"C:\Users\Zanga Musakuzi\Desktop\NSCCU DATA ANALYSIS\PROVINCIAL  tracer SUBMISSION\province submissions\march province submission\week 2\LUSAKA PROVINCE 2026 TRACER WEEKLY REPORT PROVINCES 13.03.2026.xlsx")},
            {"province": "SOUTHERN PROVINCE", "path": Path(r"C:\Users\Zanga Musakuzi\Desktop\NSCCU DATA ANALYSIS\PROVINCIAL  tracer SUBMISSION\province submissions\march province submission\week 2\SOUTHERN PROVINCE 2026 TRACER WEEKLY REPORT PROVINCES-WEEK ENDING 13.03.26 -Final.xlsx")},
        ],
        "source": "March Week 2 provincial raw submissions",
        "reportDate": "2026-03-15",
        "label": "Week 2 - 15 March 2026",
        "month": "2026-03",
        "week": "Week 2",
    },
    {
        "path": Path(r"C:\Users\Zanga Musakuzi\Desktop\NSCCU DATA ANALYSIS\PROVINCIAL  tracer SUBMISSION\tracer summery report clean data\march\22.03.26 SUMMARY REPORTS.xlsx"),
        "label": "Week 3 - 22 March 2026",
        "month": "2026-03",
        "week": "Week 3",
    },
    {
        "path": Path(r"C:\Users\Zanga Musakuzi\Desktop\NSCCU DATA ANALYSIS\PROVINCIAL  tracer SUBMISSION\tracer summery report clean data\march\DATED 29.3.26.xlsx"),
        "label": "Week 4 - 29 March 2026",
        "month": "2026-03",
        "week": "Week 4",
    },
    {
        "path": Path(r"C:\Users\Zanga Musakuzi\Desktop\NSCCU DATA ANALYSIS\PROVINCIAL  tracer SUBMISSION\tracer summery report clean data\april\TRACER SUMMARY REPORTS 4-5-2026 (1).xlsx"),
        "label": "Week 1 - 5 April 2026",
        "month": "2026-04",
        "week": "Week 1",
    },
    {
        "path": Path(r"C:\Users\Zanga Musakuzi\Desktop\NSCCU DATA ANALYSIS\PROVINCIAL  tracer SUBMISSION\tracer summery report clean data\april\TRACER SUMMARY WEEK  TWO.xlsx"),
        "label": "Week 2 - 12 April 2026",
        "month": "2026-04",
        "week": "Week 2",
    },
    {
        "path": Path(r"C:\Users\Zanga Musakuzi\Desktop\NSCCU DATA ANALYSIS\PROVINCIAL  tracer SUBMISSION\tracer summery report clean data\april\TRACER SUMMARY WEEK THREE.xlsx"),
        "label": "Week 3 - 19 April 2026",
        "month": "2026-04",
        "week": "Week 3",
    },
    {
        "path": Path(r"C:\Users\Zanga Musakuzi\Desktop\NSCCU DATA ANALYSIS\PROVINCIAL  tracer SUBMISSION\tracer summery report clean data\april\TRACER DATED 26.4.26.xlsx"),
        "label": "Week 4 - 26 April 2026",
        "month": "2026-04",
        "week": "Week 4",
    },
    {
        "path": Path(r"C:\Users\Zanga Musakuzi\Desktop\NSCCU DATA ANALYSIS\PROVINCIAL  tracer SUBMISSION\tracer summery report clean data\may\TRACER SUMMARY  WEEK ONE MAY 2026.xlsx"),
        "label": "Week 1 - 3 May 2026",
        "month": "2026-05",
        "week": "Week 1",
    },
    {
        "path": Path(r"C:\Users\Zanga Musakuzi\Desktop\NSCCU DATA ANALYSIS\PROVINCIAL  tracer SUBMISSION\analysed summery reports\may\16.05.26\tracer summary 17.05.26.xlsx"),
        "label": "Week 3 - 17 May 2026",
        "month": "2026-05",
        "week": "Week 3",
    },
    {
        "path": Path(r"C:\Users\Zanga Musakuzi\Desktop\NSCCU DATA ANALYSIS\PROVINCIAL  tracer SUBMISSION\tracer summery report clean data\may\week 3.xlsx"),
        "label": "Week 4 - 24 May 2026",
        "month": "2026-05",
        "week": "Week 4",
    },
    {
        "path": Path(r"C:\Users\Zanga Musakuzi\Desktop\NSCCU DATA ANALYSIS\PROVINCIAL  tracer SUBMISSION\tracer summery report clean data\may\WEEK 4 TRACER SUMMARY REPORT.xlsx"),
        "label": "Week 5 - 31 May 2026",
        "month": "2026-05",
        "week": "Week 5",
    },
    {
        "path": Path(r"C:\Users\Zanga Musakuzi\Desktop\NSCCU DATA ANALYSIS\PROVINCIAL  tracer SUBMISSION\tracer summery report clean data\june\07.06.2026 SUMMARY TRACE LIST.xlsx"),
        "label": "Week 1 - 7 June 2026",
        "month": "2026-06",
        "week": "Week 1",
    },
    {
        "path": Path(r"C:\Users\Zanga Musakuzi\Desktop\NSCCU DATA ANALYSIS\PROVINCIAL  tracer SUBMISSION\tracer summery report clean data\june\week 14.06.2026 tracer.xlsx"),
        "label": "Week 2 - 14 June 2026",
        "month": "2026-06",
        "week": "Week 2",
    },
    {
        "path": Path(r"C:\Users\Zanga Musakuzi\Desktop\NSCCU DATA ANALYSIS\PROVINCIAL  tracer SUBMISSION\tracer summery report clean data\june\21.06.2026 tracer.xlsx"),
        "label": "Week 3 - 21 June 2026",
        "month": "2026-06",
        "week": "Week 3",
    },
    {
        "path": JULY_WEEK4_CLEAN_WORKBOOK,
        "sheet": "Sheet1",
        "source": "28.07.26 summary reports.xlsx",
        "reportDate": "2026-07-26",
        "label": "Week 4 - 26 July 2026",
        "month": "2026-07",
        "week": "Week 4",
    },
]
OUT = Path(__file__).resolve().parents[1] / "src" / "tracerFacilityData.js"
OUT_DIR = OUT.parent
CLEAN_DATA_WORKBOOK = Path(r"C:\Users\Zanga Musakuzi\Desktop\tracer dashboard\JANUARY-DECEMBER TRACER 2026 19.07.26.xlsx")
CLEAN_DATA_SHEET = "SUMMARY SHEET"
RAW_AVAILABILITY_SOURCES = [
    {
        "reportDate": "2026-02-22",
        "province": "LUSAKA PROVINCE",
        "path": Path(r"C:\Users\Zanga Musakuzi\Desktop\NSCCU DATA ANALYSIS\PROVINCIAL  tracer SUBMISSION\province submissions\February province submission\week 3\LUSAKA PROVINCE 2026 TRACER WEEKLY REPORT PROVINCES.xlsx"),
    },
]

# Values confirmed by the programme team when a submitted workbook has been
# altered or its saved formula result is not the approved reporting value.
AUTHORITATIVE_AVAILABILITY_OVERRIDES = {
    ("2026-07-26", "CENTRAL PROVINCE", "KABWE", "LEVEL 3 HOSPITAL", "kabwe central hospital"): 0.70,
}

# Confirmed programme-team non-submissions. These level-of-care rows must not
# be interpreted as zero stock because no tracer was submitted.
CONFIRMED_NON_SUBMITTED_LEVELS = {
    ("2026-07-26", "MUCHINGA PROVINCE", "LAVUSHIMANDA", "HEALTH CENTRE"),
    ("2026-07-26", "MUCHINGA PROVINCE", "LAVUSHIMANDA", "HEALTH POST"),
    ("2026-07-26", "MUCHINGA PROVINCE", "LAVUSHIMANDA", "PRIMARY CARE - NOT SPECIFIED"),
    ("2026-07-26", "MUCHINGA PROVINCE", "NAKONDE", "HEALTH CENTRE"),
    ("2026-07-26", "MUCHINGA PROVINCE", "NAKONDE", "HEALTH POST"),
    ("2026-07-26", "MUCHINGA PROVINCE", "NAKONDE", "PRIMARY CARE - NOT SPECIFIED"),
}

# Provincial Week 4 submissions provide the verified facility-to-district
# reference used to correct copied facility labels in the clean Jan-Jun file.
RAW_FACILITY_REFERENCE_SOURCES = [
    {"province": "MUCHINGA PROVINCE", "path": Path(r"C:\Users\Zanga Musakuzi\Desktop\NSCCU DATA ANALYSIS\PROVINCIAL  tracer SUBMISSION\province submissions\june province submission\week 4\26.6.26 MUCHINGA 2026 TRACER WEEKLY REPORT.xlsx")},
    {"province": "NORTHERN PROVINCE", "path": Path(r"C:\Users\Zanga Musakuzi\Desktop\NSCCU DATA ANALYSIS\PROVINCIAL  tracer SUBMISSION\province submissions\june province submission\week 4\27.06.26 NORTHERN PROVINCE 2026 TRACER WEEKLY REPORT PROVINCES.xlsx")},
    {"province": "COPPERBELT PROVINCE", "path": Path(r"C:\Users\Zanga Musakuzi\Desktop\NSCCU DATA ANALYSIS\PROVINCIAL  tracer SUBMISSION\province submissions\june province submission\week 4\28.6.26 COPPERBELT PROVINCE  TRACER WEEKLY REPORT PROVINCES.xlsx")},
    {"province": "CENTRAL PROVINCE", "path": Path(r"C:\Users\Zanga Musakuzi\Desktop\NSCCU DATA ANALYSIS\PROVINCIAL  tracer SUBMISSION\province submissions\june province submission\week 4\28_6_2026 CENTRAL PROVINCE 2026 TRACER WEEKLY REPORT.xlsx")},
    {"province": "NORTH-WESTERN PROVINCE", "path": Path(r"C:\Users\Zanga Musakuzi\Desktop\NSCCU DATA ANALYSIS\PROVINCIAL  tracer SUBMISSION\province submissions\june province submission\week 4\28-06-2026 NORTHWESTERN TRACER WEEKLY REPORT PROVINCES.xlsx")},
    {"province": "WESTERN PROVINCE", "path": Path(r"C:\Users\Zanga Musakuzi\Desktop\NSCCU DATA ANALYSIS\PROVINCIAL  tracer SUBMISSION\province submissions\june province submission\week 4\28-06-2026 WESTERN PROVINCE 2025 TRACER WEEKLY REPORT.xlsx")},
    {"province": "EASTERN PROVINCE", "path": Path(r"C:\Users\Zanga Musakuzi\Desktop\NSCCU DATA ANALYSIS\PROVINCIAL  tracer SUBMISSION\province submissions\june province submission\week 4\28th June EASTERN PROVINCE 2026 TRACER WEEKLY REPORT PROVINCES (27).xlsx")},
    {"province": "LUAPULA PROVINCE", "path": Path(r"C:\Users\Zanga Musakuzi\Desktop\NSCCU DATA ANALYSIS\PROVINCIAL  tracer SUBMISSION\province submissions\june province submission\week 4\LUAPULA PROVINCE 2026 TRACER WEEKLY REPORT 27 6 26.xlsx")},
    {"province": "LUSAKA PROVINCE", "path": Path(r"C:\Users\Zanga Musakuzi\Desktop\NSCCU DATA ANALYSIS\PROVINCIAL  tracer SUBMISSION\province submissions\june province submission\week 4\LUSAKA PROVINCE 2026 TRACER WEEKLY REPORT PROVINCES-26.06.2026.xlsx")},
    {"province": "SOUTHERN PROVINCE", "path": Path(r"C:\Users\Zanga Musakuzi\Desktop\NSCCU DATA ANALYSIS\PROVINCIAL  tracer SUBMISSION\province submissions\june province submission\week 4\SOUTHERN PROVINCE 2026 TRACER WEEKLY REPORT PROVINCES-WEEK ENDING 26.06.26 -final.xlsx")},
    {"province": "COPPERBELT PROVINCE", "path": Path(r"C:\Users\Zanga Musakuzi\Desktop\NSCCU DATA ANALYSIS\PROVINCIAL  tracer SUBMISSION\province submissions\july\week 4\24.7.26 COPPERBELT PROVINCE  TRACER WEEKLY REPORT PROVINCES (1).xlsx")},
    {"province": "MUCHINGA PROVINCE", "path": Path(r"C:\Users\Zanga Musakuzi\Desktop\NSCCU DATA ANALYSIS\PROVINCIAL  tracer SUBMISSION\province submissions\july\week 4\25.7.26 MUCHINGA 2026 TRACER WEEKLY REPORT.xlsx")},
    {"province": "EASTERN PROVINCE", "path": Path(r"C:\Users\Zanga Musakuzi\Desktop\NSCCU DATA ANALYSIS\PROVINCIAL  tracer SUBMISSION\province submissions\july\week 4\25th July EASTERN PROVINCE 2026 TRACER WEEKLY REPORT PROVINCES (32).xlsx")},
    {"province": "NORTHERN PROVINCE", "path": Path(r"C:\Users\Zanga Musakuzi\Desktop\NSCCU DATA ANALYSIS\PROVINCIAL  tracer SUBMISSION\province submissions\july\week 4\26.07.26 NORTHERN PROVINCE 2026 TRACER WEEKLY REPORT PROVINCES.xlsx")},
    {"province": "CENTRAL PROVINCE", "path": Path(r"C:\Users\Zanga Musakuzi\Desktop\NSCCU DATA ANALYSIS\PROVINCIAL  tracer SUBMISSION\province submissions\july\week 4\26_7_2026 CENTRAL PROVINCE 2026 TRACER WEEKLY REPORT.xlsx")},
    {"province": "NORTH-WESTERN PROVINCE", "path": Path(r"C:\Users\Zanga Musakuzi\Desktop\NSCCU DATA ANALYSIS\PROVINCIAL  tracer SUBMISSION\province submissions\july\week 4\26-07-2026 NORTHWESTERN TRACER WEEKLY REPORT PROVINCES.xlsx")},
    {"province": "WESTERN PROVINCE", "path": Path(r"C:\Users\Zanga Musakuzi\Desktop\NSCCU DATA ANALYSIS\PROVINCIAL  tracer SUBMISSION\province submissions\july\week 4\26-07-2026 WESTERN PROVINCE 2025 TRACER WEEKLY REPORT.xlsx")},
    {"province": "LUAPULA PROVINCE", "path": Path(r"C:\Users\Zanga Musakuzi\Desktop\NSCCU DATA ANALYSIS\PROVINCIAL  tracer SUBMISSION\province submissions\july\week 4\LUAPULA PROVINCE 2026 TRACER WEEKLY REPORT 25 7 26.xlsx")},
    {"province": "LUSAKA PROVINCE", "path": Path(r"C:\Users\Zanga Musakuzi\Desktop\NSCCU DATA ANALYSIS\PROVINCIAL  tracer SUBMISSION\province submissions\july\week 4\LUSAKA PROVINCE 2026 TRACER WEEKLY REPORT PROVINCES-24.07.2026.xlsx")},
    {"province": "SOUTHERN PROVINCE", "path": Path(r"C:\Users\Zanga Musakuzi\Desktop\NSCCU DATA ANALYSIS\PROVINCIAL  tracer SUBMISSION\province submissions\july\week 4\SOUTHERN PROVINCE 2026 TRACER WEEKLY REPORT PROVINCES-WEEK ENDING 24.07.26 (.xlsx")},
]
RAW_FACILITY_IDENTITIES = {}
TRUSTED_RAW_FACILITY_KEYS = set()

# Authoritative corrections for named hospitals whose copied labels appear in
# unrelated province/district rows in the clean source workbook. These take
# precedence over a later raw-sheet match so one facility has one home.
VERIFIED_FACILITY_IDENTITIES = {
    "chainama hills hospital": (
        "LUSAKA PROVINCE",
        "LUSAKA",
        "MENTAL HEALTH UNITS",
        "Chainama Hills Hospital",
    ),
    "chama district hospital": (
        "MUCHINGA PROVINCE",
        "CHAMA",
        "LEVEL 1 HOSPITAL",
        "Chama District Hospital",
    ),
    "chikuni mission hospital": (
        "SOUTHERN PROVINCE",
        "MONZE",
        "LEVEL 1 HOSPITAL",
        "Chikuni Mission Hospital",
    ),
    "chipata district hospital": (
        "EASTERN PROVINCE",
        "CHIPATA",
        "LEVEL 1 HOSPITAL",
        "Chipata District Hospital",
    ),
    "chitambo district hospital": (
        "CENTRAL PROVINCE",
        "CHITAMBO",
        "LEVEL 1 HOSPITAL",
        "Chitambo District Hospital",
    ),
    "gwembe district hospital": (
        "SOUTHERN PROVINCE",
        "GWEMBE",
        "LEVEL 1 HOSPITAL",
        "Gwembe District Hospital",
    ),
    "itezhi tezhi district hospital": (
        "SOUTHERN PROVINCE",
        "ITEZHI TEZHI",
        "LEVEL 1 HOSPITAL",
        "Itezhi-tezhi District Hospital",
    ),
    "kabwe central hospital": (
        "CENTRAL PROVINCE",
        "KABWE",
        "LEVEL 3 HOSPITAL",
        "Kabwe Central Hospital",
    ),
    "national heart hospital": (
        "LUSAKA PROVINCE",
        "LUSAKA",
        "NATIONAL HEART HOSPITAL",
        "National Heart Hospital",
    ),
    "uth children": (
        "LUSAKA PROVINCE",
        "LUSAKA",
        "LEVEL 3 HOSPITAL",
        "UTH Children",
    ),
    "women and new born": (
        "LUSAKA PROVINCE",
        "LUSAKA",
        "WOMEN AND NEWBORN HOSPITAL",
        "Women and Newborn Hospital",
    ),
    "kafue gorge hospital": (
        "SOUTHERN PROVINCE",
        "CHIKANKATA",
        "LEVEL 1 HOSPITAL",
        "Kafue Gorge Hospital",
    ),
}


def clean(value):
    if value is None:
        return None
    if isinstance(value, str):
        text = " ".join(value.strip().split())
        return text or None
    return value


def normalize_header(value):
    text = (clean(value) or "").upper()
    aliases = {
        "AVERAGE AMC": "AMC",
        "AVG AMC": "AMC",
        " AMC": "AMC",
        " MOS": "MOS",
    }
    return aliases.get(text, text)


def norm_text(value):
    value = clean(value) or ""
    value = re.sub(r"[^a-z0-9]+", " ", str(value).lower())
    return " ".join(value.split())


def facility_match_key(value):
    text = norm_text(value)
    aliases = {
        "levy mwanawasa uth": "levy mwanawasa university teaching hospital",
        "chadiza district hos": "chadiza district hospital",
        "chama district hos": "chama district hospital",
        "chamboli level 1": "chamboli 1st level hospital",
        "chikuni": "chikuni mission hospital",
        "chikuni hospital": "chikuni mission hospital",
        "chipata district hos": "chipata district hospital",
        "chitambo district hos": "chitambo district hospital",
        "gwembe district hos": "gwembe district hospital",
        "itezhi tezhi district hos": "itezhi tezhi district hospital",
        "kabwe central": "kabwe central hospital",
        "kabwe central hopital": "kabwe central hospital",
        "kafue gorge": "kafue gorge hospital",
        "women and newborn": "women and new born",
        "women and newborn hospital": "women and new born",
    }
    text = aliases.get(text, text)
    district_hospital = re.fullmatch(r"(.+?)\s+(?:dh|district hos|district hosp)", text)
    if district_hospital:
        return f"{district_hospital.group(1)} district hospital"
    return text


def matches_named_district(facility_key, district):
    """True only when a district hospital is recorded against its own district."""
    match = re.fullmatch(r"(.+?)\s+district\s+hospital", facility_key)
    return bool(match and normalize_district(match.group(1)) == normalize_district(district))


def normalize_province(value):
    text = (clean(value) or "Unknown").upper()
    aliases = {
        "CENTRAL": "CENTRAL PROVINCE",
        "COPPERBELT": "COPPERBELT PROVINCE",
        "EASTERN": "EASTERN PROVINCE",
        "LUAPULA": "LUAPULA PROVINCE",
        "LUSAKA": "LUSAKA PROVINCE",
        "MUCHINGA": "MUCHINGA PROVINCE",
        "NORTHWESTERN PROVINCE": "NORTH-WESTERN PROVINCE",
        "NORTHERN": "NORTHERN PROVINCE",
        "SOUTHERN": "SOUTHERN PROVINCE",
        "WESTERN": "WESTERN PROVINCE",
    }
    return aliases.get(text, text)


def normalize_district(value):
    text = (clean(value) or "Unknown district").upper()
    text = text.replace("`", "'")
    text = re.sub(r"\s+DISTRICT$", "", text).strip()
    if text in {
        "CENTRAL PROVINCE", "COPPERBELT PROVINCE", "EASTERN PROVINCE", "LUAPULA PROVINCE",
        "LUSAKA PROVINCE", "MUCHINGA PROVINCE", "NORTH-WESTERN PROVINCE", "NORTHERN PROVINCE",
        "SOUTHERN PROVINCE", "WESTERN PROVINCE",
    }:
        return "UNKNOWN"
    aliases = {
        "SINZONGWE": "SINAZONGWE",
        "CHKANKATA": "CHIKANKATA",
        "NAWMALA": "NAMWALA",
        "MWENSE D HOSP": "MWENSE",
        "0": "UNKNOWN",
        "0.0": "UNKNOWN",
        # Loloma is a locality/ward and Loloma Mission Hospital reports under
        # Manyinga District; it is not a separate North-Western district.
        "LOLOMA": "MANYINGA",
        "UNKNOWN DISTRICT": "UNKNOWN",
    }
    return aliases.get(text, text)


# Authoritative national district roster. Raw provincial templates sometimes
# retain extra facility columns after their real district sections; enforcing
# this roster prevents facility names, comments, and copied labels from being
# counted as districts or assigned to the wrong province.
VALID_DISTRICTS_BY_PROVINCE = {
    "CENTRAL PROVINCE": {
        "CHIBOMBO", "CHISAMBA", "CHITAMBO", "KABWE", "KAPIRI MPOSHI",
        "LUANO", "MKUSHI", "MUMBWA", "NGABWE", "SERENJE", "SHIBUYUNJI",
    },
    "COPPERBELT PROVINCE": {
        "CHILILABOMBWE", "CHINGOLA", "KALULUSHI", "KITWE", "LUANSHYA",
        "LUFWANYAMA", "MASAITI", "MPONGWE", "MUFULIRA", "NDOLA",
    },
    "EASTERN PROVINCE": {
        "CHADIZA", "CHASEFU", "CHIPANGALI", "CHIPATA", "KASENENGWA",
        "KATETE", "LUMEZI", "LUNDAZI", "LUSANGAZI", "MAMBWE", "NYIMBA",
        "PETAUKE", "SINDA", "VUBWI", "LUNDAZI",
    },
    "LUAPULA PROVINCE": {
        "CHEMBE", "CHIENGE", "CHIFUNABULI", "CHIPILI", "KAWAMBWA", "LUNGA",
        "MANSA", "MILENGE", "MWANSABOMBWE", "MWENSE", "NCHELENGE", "SAMFYA",
    },
    "LUSAKA PROVINCE": {"CHILANGA", "CHONGWE", "KAFUE", "LUANGWA", "LUSAKA", "RUFUNSA"},
    "MUCHINGA PROVINCE": {
        "CHAMA", "CHINSALI", "ISOKA", "KANCHIBIYA", "LAVUSHIMANDA", "MAFINGA",
        "MPIKA", "NAKONDE", "SHIWANG'ANDU",
    },
    "NORTH-WESTERN PROVINCE": {
        "CHAVUMA", "IKELENGE", "KABOMPO", "KALUMBILA", "KASEMPA", "MANYINGA",
        "MUFUMBWE", "MUSHINDAMO", "MWINILUNGA", "SOLWEZI", "ZAMBEZI",
    },
    "NORTHERN PROVINCE": {
        "CHILUBI", "KAPUTA", "KASAMA", "LUNTE", "LUPOSOSHI", "LUWINGU", "MBALA",
        "MPOROKOSO", "MPULUNGU", "MUNGWI", "NSAMA", "SENGA",
    },
    "SOUTHERN PROVINCE": {
        "CHIKANKATA", "CHIRUNDU", "CHOMA", "GWEMBE", "ITEZHI-TEZHI", "KALOMO",
        "KAZUNGULA", "LIVINGSTONE", "MAZABUKA", "MONZE", "NAMWALA", "PEMBA",
        "SIAVONGA", "SINAZONGWE", "ZIMBA",
    },
    "WESTERN PROVINCE": {
        "KALABO", "KAOMA", "LIMULUNGA", "LUAMPA", "LUKULU", "MITETE", "MONGU",
        "MULOBEZI", "MWANDI", "NALOLO", "NKEYEMA", "SENANGA", "SESHEKE", "SHANGOMBO",
        "SIKONGO", "SIOMA",
    },
}


def canonical_district(province, district):
    district = normalize_district(district)
    aliases = {
        "KAPIRI": "KAPIRI MPOSHI",
        "ITEZHI TEZHI": "ITEZHI-TEZHI",
    }
    district = aliases.get(district, district)
    return district if district in VALID_DISTRICTS_BY_PROVINCE.get(province, set()) else None


def normalize_program(value, item=None):
    text = str(clean(value) or "Unknown programme").upper()
    item_text = str(clean(item) or "").upper()
    compact = re.sub(r"[^A-Z0-9]+", "", text)
    aliases = {
        "TB0MDR": "TB-MDR",
        "TBMDR": "TB-MDR",
        "MDRTB": "TB-MDR",
        "TB0DS": "TB-DS",
        "TBDS": "TB-DS",
        "DSTB": "TB-DS",
    }
    if compact in aliases:
        return aliases[compact]

    antiretroviral_terms = (
        "ABACAVIR", "LAMIVUDINE", "DOLUTEGRAVIR", "TENOFOVIR", "EFAVIRENZ",
        "NEVIRAPINE", "LOPINAVIR", "RITONAVIR", "ZIDOVUDINE",
    )
    if any(term in text or term in item_text for term in antiretroviral_terms):
        return "ART"
    if compact in {"", "REF", "NA", "NAN", "NONE", "UNKNOWNPROGRAMME"}:
        return "Unknown programme"
    return text


def num(value):
    if value is None:
        return None
    if isinstance(value, str) and value.strip() in {"", "-", "N/A", "#N/A"}:
        return None
    try:
        value = float(value)
    except (TypeError, ValueError):
        return None
    return None if math.isnan(value) else value


def availability_value(value):
    value = num(value)
    if value is None:
        return None
    if value > 1:
        value = value / 100 if value <= 100 else 1
    return max(0, min(value, 1))


def status_from_mos(mos):
    if mos is None:
        return "dataGap"
    if mos <= 0.1:
        return "stockout"
    if mos < 1:
        return "nearCritical"
    if mos > 12:
        return "overstock"
    if mos > 4:
        return "abovePlan"
    if mos >= 2:
        return "accordingToPlan"
    return "understocked"


def make_bucket():
    return {
        "rows": 0, "availableRows": 0, "availabilitySum": 0.0,
        "availabilityCount": 0, "mosSum": 0.0, "mosCount": 0,
        "stockout": 0, "nearCritical": 0, "understocked": 0,
        "accordingToPlan": 0, "abovePlan": 0, "overstock": 0,
        "dataGap": 0, "quantity": 0.0, "amc": 0.0,
    }


def add(bucket, row):
    mos = num(row.get("MOS"))
    availability = availability_value(row.get("AVAILABILITY"))
    bucket["rows"] += 1
    bucket["quantity"] += num(row.get("QUANTITY")) or 0
    bucket["amc"] += num(row.get("AMC")) or 0
    if availability is not None:
        bucket["availabilitySum"] += availability
        bucket["availabilityCount"] += 1
        if availability > 0:
            bucket["availableRows"] += 1
    if mos is not None:
        bucket["mosSum"] += mos
        bucket["mosCount"] += 1
    bucket[status_from_mos(mos)] += 1


def canonical_facility_level(value):
    """Return one controlled level-of-care label for every source variation."""
    text = (clean(value) or "").upper().replace("&", " ")
    if not text:
        return "PRIMARY CARE - NOT SPECIFIED"
    if "HEALTH CENTER/ HEALTH POST" in text or "HEALTH CENTRE/ HEALTH POST" in text:
        return "PRIMARY CARE - NOT SPECIFIED"
    if "HEALTH POST" in text:
        return "HEALTH POST"
    if "HEALTH CENTRE" in text or "HEALTH CENTER" in text:
        return "HEALTH CENTRE"
    if "CANCER" in text:
        return "CANCER DISEASES HOSPITAL"
    if "HEART" in text or "CARDIAC" in text:
        return "NATIONAL HEART HOSPITAL"
    if "WOMEN" in text or "NEW BORN" in text or "NEWBORN" in text or "WNB" in text:
        return "WOMEN AND NEWBORN HOSPITAL"
    if "RENAL" in text:
        return "RENAL UNITS"
    if "MENTAL" in text or "PSYCH" in text:
        return "MENTAL HEALTH UNITS"
    if "OPTH" in text or "OPHTH" in text or "EYE" in text:
        return "EYE/OPHTHALMOLOGY HOSPITAL"
    if "TB" in text or "MDR" in text:
        return "TB-DS/TB-MDR UNITS"
    if "LEVEL 1" in text or "L-1" in text or "DISTRICT LEVEL" in text:
        return "LEVEL 1 HOSPITAL"
    if "LEVEL 2" in text or "L-2" in text or "GENERAL HOSPITAL" in text:
        return "LEVEL 2/GENERAL HOSPITAL"
    if "LEVEL 3" in text or "L-3" in text or "TERTIARY" in text or "PAEDIATRIC" in text or "ADULT HOSPITAL" in text:
        return "LEVEL 3 HOSPITAL"
    if "PRIMARY FACILITY" in text:
        return "PRIMARY CARE - NOT SPECIFIED"
    return "PRIMARY CARE - NOT SPECIFIED"


def canonical_facility_level_for_facility(level, facility_name):
    """Apply known national facility classifications after source cleanup."""
    facility_key = facility_match_key(facility_name)
    if facility_key == "kafue general hospital":
        return "LEVEL 2/GENERAL HOSPITAL"
    return canonical_facility_level(level)


def canonical_facility_identity(province, district, facility_level, facility_name):
    """Apply verified facility identities and reject spreadsheet formula artefacts.

    The clean source workbook carries a small number of copied facility labels in
    unrelated province/district rows. These must not become reporting gaps. The
    identities below are verified from the provincial tracer submissions.
    """
    facility_text = str(clean(facility_name) or "").strip()
    facility_key = facility_match_key(facility_text)
    if not facility_text or facility_text.startswith("=") or facility_key in {"program a3", "ref"}:
        return None

    # Chama is a Muchinga district. Older copied source rows still label it as
    # Eastern Province, which would incorrectly create a 117th district.
    if district == "CHAMA":
        province = "MUCHINGA PROVINCE"

    verified_identity = VERIFIED_FACILITY_IDENTITIES.get(facility_key)
    if verified_identity:
        return verified_identity

    if "arthur" in facility_key and ("davison" in facility_key or "davidson" in facility_key):
        return (
            "COPPERBELT PROVINCE",
            "NDOLA",
            "LEVEL 3 HOSPITAL",
            "Arthur Children's Davison Hospital",
        )

    verified_identity = RAW_FACILITY_IDENTITIES.get(facility_key)
    if verified_identity:
        verified_province, verified_district, verified_level = verified_identity
        return verified_province, verified_district, verified_level, facility_text

    province = normalize_province(province)
    district = canonical_district(province, district)
    if district is None:
        return None
    return province, district, facility_level, facility_text


def raw_facility_level(sheet_name, sheet_title):
    title = f"{sheet_name} {sheet_title or ''}".upper()
    if sheet_name.upper() == "HP" or "HEALTH POST" in title:
        return canonical_facility_level("HEALTH POST")
    if sheet_name.upper() == "HC" or "HEALTH CENTRE" in title:
        return canonical_facility_level("HEALTH CENTRE")
    if "L-1" in title or "LEVEL 1" in title:
        return canonical_facility_level("DISTRICT LEVEL 1 HOSPITALS")
    if "L-2" in title or "LEVEL 2" in title:
        return canonical_facility_level("LEVEL 2 HOSPITAL")
    if "L-3" in title or "LEVEL 3" in title:
        return canonical_facility_level(sheet_name.upper())
    if "TB" in title:
        return canonical_facility_level("TB UNITS")
    if "EYE" in title or "OPTH" in title:
        return canonical_facility_level("OPTHAMOLOGY UNITS")
    if "RENAL" in title:
        return canonical_facility_level("RENAL UNITS")
    if "CANCER" in title:
        return canonical_facility_level("CANCER DISEASES UNITS")
    if "MENTAL" in title:
        return canonical_facility_level("MENTAL HEALTH UNITS")
    return canonical_facility_level(sheet_name.upper())


def normalize_raw_facility_level(sheet_name, sheet_title):
    return canonical_facility_level(raw_facility_level(sheet_name, sheet_title))


def raw_sheet_facility_blocks(ws):
    """Return quantity, district and facility columns for each reporting block."""
    sheet_name = ws.title.strip().upper()
    if sheet_name in {"HEART", "CANCER"}:
        return [(4, 3, 4)]
    if sheet_name in {"PSYCH", "WNB", "W & NB"}:
        return [(4, 4, 5)]
    return [(start_col, start_col, start_col + 1) for start_col in range(5, (ws.max_column or 0) + 1, 4)]


def load_raw_facility_identities():
    """Build a verified named-hospital roster from the provincial tracer files.

    Only unique named facilities are used. Aggregate health-post/health-centre
    rows are intentionally excluded because their submitted files do not name
    individual facilities.
    """
    candidates = defaultdict(set)
    skip = {"SITUATION BOARD", "SUMMARY", "COMMENTS", "COMMENT", "RECOMMENDATIONS", "RECOMEDATIONS"}
    for source in RAW_FACILITY_REFERENCE_SOURCES:
        path = source["path"]
        if not path.exists() or path.name.startswith("~$"):
            continue
        # Some provincial templates have an invalid saved worksheet dimension.
        # A full read is required here; read-only mode reports no max columns and
        # silently skips valid reporting-unit headers such as Chavuma DH.
        wb = openpyxl.load_workbook(path, read_only=False, data_only=True)
        province = normalize_province(source["province"])
        for ws in wb.worksheets:
            sheet_name = ws.title.strip()
            if any(token in sheet_name.upper() for token in skip):
                continue
            if (ws.max_row or 0) < 4 or (ws.max_column or 0) < 7:
                continue
            item_header = str(ws.cell(3, 2).value or "").upper()
            if "DESCRIPTION" not in item_header and "PRODUCT" not in item_header:
                continue
            raw_level = normalize_raw_facility_level(sheet_name, ws.cell(1, 1).value)
            if raw_level in {"HEALTH POST", "HEALTH CENTRE", "PRIMARY CARE - NOT SPECIFIED"}:
                continue
            for start_col, district_col, facility_col in raw_sheet_facility_blocks(ws):
                district = clean(ws.cell(2, district_col).value)
                facility = clean(ws.cell(2, facility_col).value)
                facility_key = facility_match_key(facility)
                verified = VERIFIED_FACILITY_IDENTITIES.get(facility_key)
                if not district and verified:
                    district = verified[1]
                if not district or not facility_key or facility_key in {"all", "hc hp", "ref"} or str(facility).startswith("="):
                    continue
                district = canonical_district(province, district)
                if district is None:
                    continue
                facility_level = canonical_facility_level_for_facility(raw_level, facility)
                candidates[facility_key].add((province, district, facility_level))
        wb.close()
    identities = {}
    trusted_keys = set()
    for key, values in candidates.items():
        matching_districts = {value for value in values if matches_named_district(key, value[1])}
        if len(matching_districts) == 1:
            identities[key] = next(iter(matching_districts))
            trusted_keys.add(key)
        elif len(values) == 1 and key in VERIFIED_FACILITY_IDENTITIES:
            # A curated named facility is allowed to use its provincial raw
            # submission even when its title does not include the district.
            identities[key] = next(iter(values))
            trusted_keys.add(key)
    TRUSTED_RAW_FACILITY_KEYS.clear()
    TRUSTED_RAW_FACILITY_KEYS.update(trusted_keys)
    return identities


def load_raw_availability_overrides():
    overrides = {}
    skip = {"SITUATION BOARD", "SUMMARY", "SUMMARY SHEET", "COMMENTS", "COMMENT", "RECOMMENDATIONS", "RECOMEDATIONS", "PROGRAM"}
    for source in RAW_AVAILABILITY_SOURCES:
        if not source["path"].exists():
            continue
        wb = openpyxl.load_workbook(source["path"], read_only=False, data_only=True)
        province = normalize_province(source["province"])
        report_date = source["reportDate"]
        for ws in wb.worksheets:
            sheet_name = ws.title.strip()
            if any(token in sheet_name.upper() for token in skip):
                continue
            footer_row = None
            for row_index in range(1, ws.max_row + 1):
                label = str(ws.cell(row_index, 2).value or "").strip().upper()
                if "PERCENTAGE AVAILABILITY" in label:
                    footer_row = row_index
                    break
            if not footer_row:
                continue
            for start_col in range(5, ws.max_column + 1, 4):
                district = clean(ws.cell(2, start_col).value)
                facility = clean(ws.cell(2, start_col + 1).value)
                availability = availability_value(ws.cell(footer_row, start_col).value)
                if not district or not facility or availability is None:
                    continue
                facility_level = canonical_facility_level_for_facility(
                    normalize_raw_facility_level(sheet_name, ws.cell(1, 1).value),
                    facility,
                )
                key = (report_date, province, normalize_district(district), facility_level, facility_match_key(facility))
                overrides[key] = availability
        wb.close()
    return overrides


def iter_raw_matrix_rows(source, report_date):
    workbook_path = source["path"]
    province = normalize_province(source["province"])
    wb = openpyxl.load_workbook(workbook_path, read_only=False, data_only=True)
    skip = {"SITUATION BOARD", "SUMMARY", "COMMENTS", "COMMENT", "RECOMMENDATIONS", "RECOMEDATIONS"}
    for ws in wb.worksheets:
        sheet_name = ws.title.strip()
        if any(token in sheet_name.upper() for token in skip):
            continue
        if ws.max_row < 4 or ws.max_column < 7:
            continue
        item_header = str(ws.cell(3, 2).value or "").upper()
        if "DESCRIPTION" not in item_header and "PRODUCT" not in item_header:
            continue
        source_facility_level = raw_facility_level(sheet_name, ws.cell(1, 1).value)
        for start_col, district_col, facility_col in raw_sheet_facility_blocks(ws):
            district = clean(ws.cell(2, district_col).value)
            facility = clean(ws.cell(2, facility_col).value)
            verified = VERIFIED_FACILITY_IDENTITIES.get(facility_match_key(facility))
            if not district and verified:
                district = verified[1]
            if not district or not facility:
                continue
            district = canonical_district(province, district)
            if district is None:
                continue
            facility_level = canonical_facility_level_for_facility(source_facility_level, facility)
            if facility_level in {"HEALTH POST", "HEALTH CENTRE"}:
                facility_name = f"{district} {facility_level.title()} facilities"
                is_aggregate = True
            else:
                facility_name = str(facility or district).strip()
                is_aggregate = False
            identity = canonical_facility_identity(province, district, facility_level, facility_name)
            if identity is None:
                continue
            province, district, facility_level, facility_name = identity
            for row_index in range(4, ws.max_row + 1):
                item = clean(ws.cell(row_index, 2).value)
                if not item:
                    continue
                if "PERCENTAGE AVAILABILITY" in str(item).upper():
                    continue
                quantity = num(ws.cell(row_index, start_col).value)
                amc = num(ws.cell(row_index, start_col + 1).value)
                mos = num(ws.cell(row_index, start_col + 2).value)
                if quantity is None and amc is None and mos is None:
                    continue
                yield {
                    "DATE": report_date,
                    "NATION": "ZAMBIA",
                    "PROVINCE": province,
                    "DISTRICT": district,
                    "FACILITY LEVEL": facility_level,
                    "FACILITY NAME": facility_name,
                    "PROGRAM": normalize_program(facility_level),
                    "DESCRIPTION OF ITEM": item,
                    "UNIT": clean(ws.cell(row_index, 3).value),
                    "QUANTITY": quantity,
                    "AMC": amc,
                    "MOS": mos,
                    "AVAILABILITY": 1 if (quantity or 0) > 0 else 0,
                    "_RAW_AGGREGATE": is_aggregate,
                }
    wb.close()


def date_id(value):
    if hasattr(value, "strftime"):
        return value.strftime("%Y-%m-%d")
    return str(value)[:10]


def week_label(value):
    report_id = date_id(value)
    week_overrides = {
        "2026-02-22": "Week 3",
        "2026-02-28": "Week 4",
        "2026-03-08": "Week 1",
        "2026-03-15": "Week 2",
        "2026-03-22": "Week 3",
        "2026-03-29": "Week 4",
    }
    if report_id in week_overrides:
        return week_overrides[report_id]
    if hasattr(value, "strftime"):
        return f"Week {(value.day - 1) // 7 + 1}"
    try:
        day = int(str(value)[8:10])
    except (TypeError, ValueError):
        day = 1
    return f"Week {(day - 1) // 7 + 1}"


def month_id(value):
    if hasattr(value, "strftime"):
        return value.strftime("%Y-%m")
    return str(value)[:7]


def period_label(value):
    if hasattr(value, "strftime"):
        return f"{week_label(value)} - {value.day} {value.strftime('%B %Y')}"
    return f"{week_label(value)} - {date_id(value)}"


def corrected_value(value, fallback):
    text = clean(value)
    if not text or str(text).upper() in {"NOT FOUND IN MASTER LIST", "#N/A", "#REF!"}:
        return clean(fallback)
    return text


def clean_facility_level(level, corrected_level):
    return canonical_facility_level(corrected_value(corrected_level, level))


def load_clean_workbook_configs(workbook_path=CLEAN_DATA_WORKBOOK, sheet_name=CLEAN_DATA_SHEET, source_label=None):
    wb = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    ws = wb[sheet_name]
    rows_by_date = defaultdict(list)
    date_values = {}
    source = source_label or workbook_path.name

    for values in ws.iter_rows(min_row=2, values_only=True):
        if not any(values):
            continue
        report_date = values[0]
        if not report_date:
            continue
        report_id = date_id(report_date)
        date_values[report_id] = report_date
        province = normalize_province(values[2])
        district = normalize_district(values[3])
        # Do not carry spreadsheet placeholder values such as a numeric zero
        # into the national reporting geography.
        if district == "UNKNOWN":
            continue
        original_level = clean(values[4])
        original_facility = clean(values[5])
        original_item = clean(values[7])
        corrected_facility = corrected_value(values[17] if len(values) > 17 else None, original_facility)
        if str(corrected_facility or "").strip().upper() == "HC/HP" and str(original_facility or "").strip().upper() not in {"", "ALL"}:
            corrected_facility = original_facility
        corrected_item = corrected_value(values[18] if len(values) > 18 else None, original_item)
        programme = normalize_program(values[6], corrected_item or original_item)
        facility_level = clean_facility_level(original_level, values[19] if len(values) > 19 else None)
        is_aggregate = (
            str(original_facility or "").strip().upper() == "ALL"
            or str(corrected_facility or "").strip().upper() == "ALL"
            or (
                str(corrected_facility or "").strip().upper() == "HC/HP"
                and facility_level in {"HEALTH POST", "HEALTH CENTRE", "PRIMARY CARE - NOT SPECIFIED"}
            )
        )
        facility_name = corrected_facility or original_facility or "Unknown reporting unit"
        facility_level = canonical_facility_level_for_facility(facility_level, facility_name)
        if is_aggregate and facility_level in {"HEALTH POST", "HEALTH CENTRE"}:
            facility_name = f"{district} {facility_level.title()} facilities"

        identity = canonical_facility_identity(province, district, facility_level, facility_name)
        if identity is None:
            continue
        province, district, facility_level, facility_name = identity

        quantity = num(values[9] if len(values) > 9 else None)
        rows_by_date[report_id].append({
            "DATE": report_id,
            "NATION": "ZAMBIA",
            "PROVINCE": province,
            "DISTRICT": district,
            "FACILITY LEVEL": facility_level,
            "FACILITY NAME": facility_name,
            "PROGRAM": programme or "Unknown programme",
            "DESCRIPTION OF ITEM": corrected_item or original_item or "Unknown commodity",
            "UNIT": clean(values[8] if len(values) > 8 else None),
            "QUANTITY": quantity,
            "AMC": num(values[10] if len(values) > 10 else None),
            "MOS": num(values[11] if len(values) > 11 else None),
            "AVAILABILITY": 1 if (quantity or 0) > 0 else 0,
            "COMMENT": clean(values[13] if len(values) > 13 else None),
            "_RAW_AGGREGATE": is_aggregate,
        })
    wb.close()

    configs = []
    for report_id, rows in rows_by_date.items():
        raw_date = date_values[report_id]
        configs.append({
            "rows": rows,
            "reportDate": report_id,
            "label": period_label(raw_date),
            "month": month_id(raw_date),
            "week": week_label(raw_date),
            "source": source,
        })
    return configs


def finalize(name, bucket, extra=None):
    rows = bucket["rows"]
    result = {
        "name": name,
        "rows": rows,
        "availability": round(bucket["availabilitySum"] / bucket["availabilityCount"], 4) if bucket["availabilityCount"] else 0,
        "mos": round(bucket["mosSum"] / bucket["mosCount"], 2) if bucket["mosCount"] else None,
        "stockout": bucket["stockout"],
        "nearCritical": bucket["nearCritical"],
        "understocked": bucket["understocked"],
        "accordingToPlan": bucket["accordingToPlan"],
        "abovePlan": bucket["abovePlan"],
        "overstock": bucket["overstock"],
        "dataGap": bucket["dataGap"],
        "quantity": round(bucket["quantity"], 2),
        "amc": round(bucket["amc"], 2),
        "riskRows": bucket["stockout"] + bucket["nearCritical"] + bucket["understocked"],
        "stockoutRate": round(bucket["stockout"] / rows, 4) if rows else 0,
    }
    if extra:
        result.update(extra)
    return result


def summarize(config):
    workbook_path = config.get("path")
    availability_overrides = config.get("availabilityOverrides", {})
    if config.get("rows") is not None:
        rows_iter = iter(config["rows"])
        source_name = config.get("source", "Clean tracer dataset")
        wb = None
    elif config.get("rawSources"):
        rows_iter = (
            row
            for source in config["rawSources"]
            for row in iter_raw_matrix_rows(source, config["reportDate"])
        )
        source_name = config.get("source", "Provincial raw submissions")
        wb = None
    else:
        wb = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
        ws = wb[config.get("sheet") or wb.sheetnames[0]]
        headers = [normalize_header(cell) for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
        rows_iter = (
            {key: clean(value) for key, value in zip(headers, row_values)}
            for row_values in ws.iter_rows(min_row=2, values_only=True)
        )
        source_name = workbook_path.name

    national = make_bucket()
    by_province = defaultdict(make_bucket)
    by_district = defaultdict(make_bucket)
    by_facility_level = defaultdict(make_bucket)
    by_facility = defaultdict(make_bucket)
    by_program = defaultdict(make_bucket)
    by_program_scope = defaultdict(make_bucket)
    by_item = defaultdict(make_bucket)
    commodity_facility_rows = []
    facility_items = defaultdict(lambda: {"stockout": [], "lowStock": [], "accordingToPlan": [], "overstock": []})
    facility_is_aggregate = {}
    comments = []
    province_names, district_units, facility_units, item_names, program_names = set(), set(), set(), set(), set()
    report_date = None

    for row in rows_iter:
        province = normalize_province(row.get("PROVINCE"))
        district = normalize_district(row.get("DISTRICT"))
        facility = clean(row.get("FACILITY NAME")) or "Unknown reporting unit"
        facility_level = canonical_facility_level(row.get("FACILITY LEVEL"))
        identity = canonical_facility_identity(province, district, facility_level, facility)
        if identity is None:
            continue
        province, district, facility_level, facility = identity
        row_report_date = config.get("reportDate") or date_id(row.get("DATE"))
        if (row_report_date, province, district, facility_level) in CONFIRMED_NON_SUBMITTED_LEVELS:
            continue
        if facility_level in {"NATIONAL HEART HOSPITAL", "WOMEN AND NEWBORN HOSPITAL"} and province != "LUSAKA PROVINCE":
            continue
        item = clean(row.get("DESCRIPTION OF ITEM")) or "Unknown commodity"
        program = normalize_program(row.get("PROGRAM"), item)
        cancer_scope = (
            program == "CANCER"
            or "CANCER" in facility_level
            or "CANCER" in facility.upper()
        )
        # Cancer indicators in the tracer are reported from the Cancer Diseases
        # Hospital in Lusaka. Exclude unrelated provincial cancer labels so every
        # tracer tab uses the same designated reporting source.
        if cancer_scope and not (
            province == "LUSAKA PROVINCE"
            and ("CANCER" in facility_level or "CANCER" in facility.upper())
        ):
            continue
        if cancer_scope:
            facility = "CANCER DISEASES HOSPITAL"
            facility_level = "CANCER DISEASES HOSPITAL"
            # Cancer performance is defined from the designated Cancer Diseases
            # Hospital tracer submission, even where individual rows use a
            # generic source-programme label.
            program = "CANCER"
        if report_date is None:
            report_date = row.get("DATE")

        province_names.add(province)
        if district != "UNKNOWN":
            district_units.add((province, district))
        facility_units.add((province, district, facility_level, facility))
        item_names.add(item)
        program_names.add(program)

        for bucket in (
            national, by_province[province], by_district[(province, district)],
            by_facility_level[facility_level],
            by_facility[(province, district, facility_level, facility)],
            by_program[program],
            by_program_scope[(province, facility_level, program)],
            by_item[item],
        ):
            add(bucket, row)

        # Compact facility-by-commodity rows support the dashboard drill-down
        # without repeating full object keys for every submitted tracer record.
        commodity_facility_rows.append((
            province,
            district,
            facility_level,
            facility,
            item,
            program,
            num(row.get("QUANTITY")),
            num(row.get("AMC")),
            num(row.get("MOS")),
        ))

        mos = num(row.get("MOS"))
        alert_item = {
            "item": item,
            "program": program,
            "mos": round(mos, 2) if mos is not None else None,
            "quantity": round(num(row.get("QUANTITY")) or 0, 2),
            "amc": round(num(row.get("AMC")) or 0, 2),
        }
        key = (province, district, facility_level, facility)
        facility_is_aggregate[key] = bool(row.get("_RAW_AGGREGATE")) or facility.upper() == "ALL"
        if mos is not None and mos <= 0.1:
            facility_items[key]["stockout"].append(alert_item)
        elif mos is not None and mos < 2:
            facility_items[key]["lowStock"].append(alert_item)
        elif mos is not None and mos <= 4:
            facility_items[key]["accordingToPlan"].append(alert_item)
        elif mos is not None:
            facility_items[key]["overstock"].append(alert_item)

        note = clean(row.get("COMMENT"))
        if note and note not in {"#NAME?", "STOCKED ACCORDING TO PLAN", "UNDERSTOCKED", "OVERSTOCKED"}:
            comments.append({"province": province, "note": note})

    provinces = [finalize(name, bucket) for name, bucket in by_province.items()]
    provinces.sort(key=lambda item: (item["availability"], -item["riskRows"]))
    districts = [
        finalize(district, bucket, {"province": province})
        for (province, district), bucket in by_district.items()
    ]
    districts.sort(key=lambda item: (item["availability"], -item["riskRows"]))
    facility_levels = [finalize(name, bucket) for name, bucket in by_facility_level.items()]
    facility_levels.sort(key=lambda item: (item["availability"], -item["riskRows"]))

    facilities = []
    for key, bucket in by_facility.items():
        province, district, facility_level, facility = key
        alerts = facility_items[key]
        according_to_plan_items = sorted(
            alerts["accordingToPlan"],
            key=lambda item: (item["mos"] if item["mos"] is not None else 99, str(item["program"]), str(item["item"])),
        )
        overstock_items = sorted(
            alerts["overstock"],
            key=lambda item: (-(item["mos"] if item["mos"] is not None else 0), str(item["program"]), str(item["item"])),
        )
        facility_result = finalize(facility, bucket, {
            "province": province,
            "district": district,
            "facilityLevel": facility_level,
            "isAggregate": facility_is_aggregate.get(key, facility.upper() == "ALL"),
            "stockoutItems": sorted(alerts["stockout"], key=lambda item: (str(item["program"]), str(item["item"])))[:60],
            "lowStockItems": sorted(alerts["lowStock"], key=lambda item: (item["mos"] if item["mos"] is not None else 99, str(item["program"]), str(item["item"])))[:60],
            "accordingToPlanItems": according_to_plan_items[:5],
            "overstockItems": overstock_items[:5],
            "stockoutItemCount": len(alerts["stockout"]),
            "lowStockItemCount": len(alerts["lowStock"]),
            "accordingToPlanItemCount": len(alerts["accordingToPlan"]),
            "overstockItemCount": len(alerts["overstock"]),
        })
        override_key = (config.get("reportDate"), province, district, facility_level, facility_match_key(facility))
        if override_key in availability_overrides:
            facility_result["availability"] = round(availability_overrides[override_key], 4)
        facilities.append(facility_result)
    facilities.sort(key=lambda item: (-item["stockoutItemCount"], -item["lowStockItemCount"], item["availability"], item["province"], item["district"], item["name"]))

    programs = [finalize(name, bucket) for name, bucket in by_program.items()]
    programs.sort(key=lambda item: (item["availability"], -item["riskRows"]))
    program_scopes = [
        finalize(program, bucket, {
            "province": province,
            "facilityLevel": facility_level,
        })
        for (province, facility_level, program), bucket in by_program_scope.items()
    ]
    program_scopes.sort(key=lambda item: (item["province"], item["facilityLevel"], item["availability"], item["name"]))
    items = [finalize(name, bucket, {"normalized": norm_text(name)}) for name, bucket in by_item.items()]
    items.sort(key=lambda item: (-item["riskRows"], item["availability"], str(item["name"])))

    if config.get("reportDate"):
        report_date = config["reportDate"]
    elif hasattr(report_date, "strftime"):
        report_date = report_date.strftime("%Y-%m-%d")

    compact_keys = ["provinces", "districts", "levels", "facilities", "items", "programmes"]
    compact_values = {key: [] for key in compact_keys}
    compact_lookup = {key: {} for key in compact_keys}

    def compact_index(key, value):
        lookup = compact_lookup[key]
        if value not in lookup:
            lookup[value] = len(compact_values[key])
            compact_values[key].append(value)
        return lookup[value]

    compact_rows = [
        [
            compact_index("provinces", province),
            compact_index("districts", district),
            compact_index("levels", facility_level),
            compact_index("facilities", facility),
            compact_index("items", item),
            compact_index("programmes", program),
            quantity,
            amc,
            mos,
        ]
        for province, district, facility_level, facility, item, program, quantity, amc, mos in commodity_facility_rows
    ]
    return {
        "id": report_date,
        "reportDate": report_date,
        "label": config["label"],
        "month": config["month"],
        "week": config["week"],
        "source": source_name,
        "counts": {
            "rows": national["rows"],
            "provinces": len(province_names),
            "districts": len(district_units),
            "facilityUnits": len(facility_units),
            "programmes": len(program_names),
            "commodities": len(item_names),
        },
        "national": finalize("Zambia", national),
        "provinces": provinces,
        "districts": districts,
        "facilities": facilities,
        "facilityLevels": facility_levels,
        "programmes": programs,
        "programmeScopes": program_scopes,
        "commodities": items,
        "commodityFacilityData": {
            "dictionaries": compact_values,
            "rows": compact_rows,
        },
        "comments": comments[:100],
    }


def reporting_rate(reported, expected):
    return round(reported / expected, 4) if expected else 0


def reporting_facility_type(facility_level):
    text = str(facility_level or "").upper()
    if "PRIMARY CARE" in text or "PRIMARY FACILITY" in text:
        return "Health Centres and Posts (combined)"
    if "HEALTH POST" in text:
        return "Health Posts"
    if "HEALTH CENTRE" in text or "HEALTH CENTER" in text:
        return "Health Centres"
    if "RENAL" in text:
        return "Renal Units"
    if "CANCER" in text:
        return "Cancer Units"
    if "MENTAL" in text:
        return "Mental Health Units"
    if "OPTH" in text or "OPHTH" in text:
        return "Ophthalmology Units"
    if "TB" in text or "DS-TB" in text or "MDR" in text:
        return "TB-DS/MDR Units"
    if "HEART" in text:
        return "Heart Hospital"
    if "WOMEN" in text or "NEW BORN" in text:
        return "Women and Newborn Hospital"
    if "PAEDIATRIC" in text:
        return "Paediatric Hospital"
    if "LEVEL 2" in text or "GENERAL HOSPITAL" in text:
        return "Level 2 Hospitals"
    if "LEVEL 1" in text or "DISTRICT" in text:
        return "Level 1 Hospitals"
    if "LEVEL 3" in text or "TERTIARY" in text or "SPECIAL" in text:
        return "Level 3/Specialised Hospitals"
    return "Other reporting units"


def build_reporting_quality(periods, expected_districts, expected_facilities):
    primary_care_types = {"Health Centres", "Health Posts"}
    combined_primary_care_type = "Health Centres and Posts (combined)"
    province_names = sorted({province for province, _district in expected_districts})
    district_names = sorted(expected_districts)
    expected_level_reports = {
        (province, district, reporting_facility_type(facility_level))
        for province, district, facility_level, _facility in expected_facilities
    }
    expected_named_reports = {
        (province, district, facility_level, facility)
        for province, district, facility_level, facility in expected_facilities
        if district != "UNKNOWN" and (
            facility_match_key(facility) in VERIFIED_FACILITY_IDENTITIES
            or facility_match_key(facility) in TRUSTED_RAW_FACILITY_KEYS
        )
    }

    for period in periods:
        present_districts = {(district["province"], district["name"]) for district in period["districts"]}
        present_facilities = {
            (facility["province"], facility["district"], facility["facilityLevel"], facility["name"])
            for facility in period["facilities"]
        }
        present_level_reports = {
            (province, district, reporting_facility_type(facility_level))
            for province, district, facility_level, _facility in present_facilities
            if district != "UNKNOWN"
        }
        submitted_dho_districts = set()
        partial_dho_districts = set()
        hospital_only_districts = set()
        for province, district in expected_districts:
            reported_types = {
                facility_type
                for report_province, report_district, facility_type in present_level_reports
                if report_province == province and report_district == district
            }
            combined_reported = combined_primary_care_type in reported_types
            split_primary_reports = primary_care_types & reported_types
            if combined_reported or split_primary_reports == primary_care_types:
                submitted_dho_districts.add((province, district))
            elif split_primary_reports:
                partial_dho_districts.add((province, district))
            elif reported_types:
                hospital_only_districts.add((province, district))
        missing_districts = sorted(expected_districts - submitted_dho_districts)
        missing_level_reports = sorted(expected_level_reports - present_level_reports)

        province_rows = []
        for province in province_names:
            expected_province_reports = {item for item in expected_level_reports if item[0] == province}
            reported_province_reports = expected_province_reports & present_level_reports
            expected_province_districts = {item for item in expected_districts if item[0] == province}
            reported_province_districts = expected_province_districts & submitted_dho_districts
            expected = len(expected_province_reports)
            reported = len(reported_province_reports)
            province_rows.append({
                "name": province,
                "districts": len(reported_province_districts),
                "expectedDistricts": len(expected_province_districts),
                "expected": expected,
                "reported": reported,
                "missing": max(expected - reported, 0),
                "rate": reporting_rate(reported, expected),
            })

        district_rows = []
        for province, district in district_names:
            expected_district_reports = {item for item in expected_level_reports if item[0] == province and item[1] == district}
            reported_district_reports = expected_district_reports & present_level_reports
            reported_types = {item[2] for item in reported_district_reports}
            combined_reported = combined_primary_care_type in reported_types
            health_centre_reported = "Health Centres" in reported_types
            health_post_reported = "Health Posts" in reported_types
            submitted = combined_reported or (health_centre_reported and health_post_reported)
            partial = not combined_reported and health_centre_reported != health_post_reported
            expected = 1
            reported = 1 if submitted else 0
            district_rows.append({
                "province": province,
                "name": district,
                "expected": expected,
                "reported": reported,
                "missing": max(expected - reported, 0),
                "rate": reporting_rate(reported, expected),
                "submitted": submitted,
                "partial": partial,
                "hospitalOnly": not submitted and not partial and bool(reported_types),
                "healthCentreReported": health_centre_reported or combined_reported,
                "healthPostReported": health_post_reported or combined_reported,
            })

        type_rows = []
        for province, district, facility_type in sorted(expected_level_reports):
            reported = 1 if (province, district, facility_type) in present_level_reports else 0
            type_rows.append({
                "province": province,
                "district": district,
                "type": facility_type,
                "expected": 1,
                "reported": reported,
                "missing": 1 - reported,
                "rate": reporting_rate(reported, 1),
            })

        facility_rows = []
        for province, district, facility_level, facility in sorted(expected_named_reports):
            reported = 1 if (province, district, facility_level, facility) in present_facilities else 0
            facility_rows.append({
                "province": province,
                "district": district,
                "facilityLevel": facility_level,
                "facilityType": reporting_facility_type(facility_level),
                "name": facility,
                "expected": 1,
                "reported": reported,
                "missing": 1 - reported,
                "rate": reporting_rate(reported, 1),
            })

        province_rows.sort(key=lambda item: (item["rate"], -item["missing"], item["name"]))
        district_rows.sort(key=lambda item: (item["rate"], -item["missing"], item["province"], item["name"]))
        type_rows.sort(key=lambda item: (item["province"], item["district"], item["type"]))
        facility_rows.sort(key=lambda item: (item["province"], item["district"], item["facilityLevel"], item["name"]))

        period["counts"]["districtSubmissionFootprint"] = period["counts"]["districts"]
        period["counts"]["districts"] = len(submitted_dho_districts)
        period["counts"]["expectedDistricts"] = len(expected_districts)
        period["counts"]["partialDistricts"] = len(partial_dho_districts)
        period["counts"]["hospitalOnlyDistricts"] = len(hospital_only_districts)
        period["counts"]["expectedFacilityUnits"] = len(expected_level_reports)
        period["counts"]["expectedLevelReports"] = len(expected_level_reports)
        period["counts"]["expectedNamedFacilityReports"] = len(expected_named_reports)
        period["counts"]["missingDistricts"] = len(missing_districts)
        period["counts"]["missingFacilityUnits"] = len(missing_level_reports)
        period["counts"]["missingLevelReports"] = len(missing_level_reports)
        period["missingDistricts"] = [
            {"province": province, "district": district}
            for province, district in missing_districts[:100]
        ]
        period["missingFacilities"] = [
            {"province": province, "district": district, "facilityLevel": facility_type, "name": facility_type}
            for province, district, facility_type in missing_level_reports[:100]
        ]
        period["dataQuality"] = {
            "districtSubmissionRule": "Both Health Centre and Health Post reporting are required; a combined primary-care source counts for both. Hospital submissions do not count.",
            "provinces": province_rows,
            "districts": district_rows,
            "facilityTypes": type_rows,
            "facilities": facility_rows,
        }


def main():
    RAW_FACILITY_IDENTITIES.clear()
    RAW_FACILITY_IDENTITIES.update(load_raw_facility_identities())
    print(f"Loaded {len(RAW_FACILITY_IDENTITIES)} verified named-facility identities from provincial submissions.")
    availability_overrides = load_raw_availability_overrides()
    availability_overrides.update(AUTHORITATIVE_AVAILABILITY_OVERRIDES)
    configs = []
    # The approved clean workbook begins on 22 February. January is retained
    # from its validated weekly summary workbook so the dashboard covers Jan-Jun.
    for config in WORKBOOKS:
        if config.get("month") == "2026-01":
            config = dict(config)
            config["availabilityOverrides"] = availability_overrides
            configs.append(config)
    for config in load_clean_workbook_configs():
        config["availabilityOverrides"] = availability_overrides
        configs.append(config)
    for config in load_clean_workbook_configs(
        JULY_WEEK4_CLEAN_WORKBOOK,
        "Sheet1",
        "28.07.26 summary reports.xlsx",
    ):
        config["availabilityOverrides"] = availability_overrides
        configs.append(config)
    for config in load_clean_workbook_configs(
        JULY_WEEK5_CLEAN_WORKBOOK,
        "Sheet1",
        "TRACER SUMMARY 02 AUAGUST 2026.xlsx",
    ):
        # Operationally this is July Week 5 even though the reporting week ends
        # on 2 August. Keep it under July so it follows the programme calendar
        # and appears after Week 4 in every dashboard tab.
        config["label"] = "Week 5 - 2 August 2026"
        config["month"] = "2026-07"
        config["week"] = "Week 5"
        config["availabilityOverrides"] = availability_overrides
        configs.append(config)
    for config in load_clean_workbook_configs(
        AUGUST_WEEK1_CLEAN_WORKBOOK,
        "Sheet1",
        "9.8.2026 tracer summary.xlsx",
    ):
        # The reporting calendar restarts in August after the 2 August
        # July close-out report, so this is August Week 1.
        config["label"] = "Week 1 - 9 August 2026"
        config["month"] = "2026-08"
        config["week"] = "Week 1"
        config["availabilityOverrides"] = availability_overrides
        configs.append(config)
    for config in load_clean_workbook_configs(
        AUGUST_WEEK2_CLEAN_WORKBOOK,
        "Sheet1",
        "tracer summary 16-08-2026.xlsx",
    ):
        config["label"] = "Week 2 - 16 August 2026"
        config["month"] = "2026-08"
        config["week"] = "Week 2"
        config["availabilityOverrides"] = availability_overrides
        configs.append(config)
    for config in load_clean_workbook_configs(
        AUGUST_WEEK3_CLEAN_WORKBOOK,
        "Sheet1",
        "23.08.2026.xlsx",
    ):
        config["label"] = "Week 3 - 23 August 2026"
        config["month"] = "2026-08"
        config["week"] = "Week 3"
        config["availabilityOverrides"] = availability_overrides
        configs.append(config)
    clean_period_ids = {config["reportDate"] for config in configs}
    # Retain the clean master as the source of record for its existing dates,
    # then add provincial submissions only for new reporting periods not yet
    # present in that master workbook.
    for config in WORKBOOKS:
        if (config.get("rawSources") or config.get("reportDate") == "2026-07-26") and config["reportDate"] not in clean_period_ids:
            config = dict(config)
            config["availabilityOverrides"] = availability_overrides
            configs.append(config)
    periods = [summarize(config) for config in configs]
    periods.sort(key=lambda item: item["reportDate"])
    expected_districts = {
        (province, district)
        for province, districts in VALID_DISTRICTS_BY_PROVINCE.items()
        for district in districts
    }
    expected_facilities = {
        (facility["province"], facility["district"], facility["facilityLevel"], facility["name"])
        for period in periods
        for facility in period["facilities"]
        if facility["district"] != "UNKNOWN"
        and not (
            period["id"] == "2026-07-26"
            and facility["facilityLevel"] == "PRIMARY CARE - NOT SPECIFIED"
        )
    }
    build_reporting_quality(periods, expected_districts, expected_facilities)
    # Keep each generated module below GitHub's 100 MB file limit while
    # preserving the complete Jan-Jun history available to the dashboard.
    period_groups = {
        "JanFeb": [period for period in periods if period["month"] in {"2026-01", "2026-02"}],
        "MarApr": [period for period in periods if period["month"] in {"2026-03", "2026-04"}],
        "MayJun": [period for period in periods if period["month"] in {"2026-05", "2026-06"}],
        "Jul": [period for period in periods if period["month"] >= "2026-07"],
    }
    module_names = []
    for suffix, group in period_groups.items():
        module_name = f"tracerFacilityData{suffix}"
        module_names.append(module_name)
        (OUT_DIR / f"{module_name}.js").write_text(
            "export const tracerReportingPeriods = " + json.dumps(group, separators=(",", ":")) + ";\n",
            encoding="utf-8",
        )

    imports = "\n".join(
        f'import {{ tracerReportingPeriods as {name} }} from "./{name}.js";'
        for name in module_names
    )
    output = (
        f"{imports}\n\n"
        + "export const availableTracerYears = [\"2024\", \"2025\", \"2026\"];\n\n"
        + "export let tracerReportingPeriods = ["
        + ", ".join(f"...{name}" for name in module_names)
        + "].sort((left, right) => left.reportDate.localeCompare(right.reportDate));\n\n"
        + "const loadedHistoricalYears = new Set();\n\n"
        + "export async function loadHistoricalTracerYear(year) {\n"
        + "  if (year === \"2026\" || loadedHistoricalYears.has(year)) return tracerReportingPeriods;\n"
        + "  if (!availableTracerYears.includes(year)) throw new Error(`No tracer data is available for ${year}.`);\n\n"
        + "  const module = await import(/* @vite-ignore */ `/historical/tracerFacilityData${year}.js`);\n"
        + "  tracerReportingPeriods = [...tracerReportingPeriods, ...module.tracerReportingPeriods]\n"
        + "    .sort((left, right) => left.reportDate.localeCompare(right.reportDate));\n"
        + "  loadedHistoricalYears.add(year);\n"
        + "  return tracerReportingPeriods;\n"
        + "}\n\nexport const tracerFacilityData = tracerReportingPeriods.at(-1);\n"
    )
    OUT.write_text(output, encoding="utf-8")
    print([(item["label"], item["counts"]) for item in periods])


if __name__ == "__main__":
    main()
